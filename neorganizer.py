from tkinter import filedialog
from openpyxl import *
from pearson import *
import tkinter as tk
import tkinter.ttk as ttk
import random

class Neorganizer:

    def __init__(self, root):
        self.community_members = []
        self.loaded = False
        self.root = root
        self.createMainScreen()
        # self.current_screen = tk.Frame(self.root)
        # self.text_frame = tk.Listbox(self.current_screen , bg='white')
        # self.text_frame.grid(row=0, column=0, sticky=tk.W)
        self.number_of_circles = 3
        self.circles = [[] for x in range(self.number_of_circles)]
        self.initCircleTextBox()

    def clearScr(self):
        self.current_screen.destroy()

    def createMainScreen(self):
        self.current_screen = tk.Frame(self.root)
        self.current_screen.pack()

    def createTextBox(self):
        self.text_frame = tk.Listbox(self.current_screen , bg='white')
        self.text_frame.pack()

    def initCircleTextBox(self):
        self.circleTextBox = [tk.Text(self.current_screen, bg='white', state=tk.DISABLED) for x in range(self.number_of_circles)]

    def loadCommunityWorkboodAndPrintContent(self):
        self.clearScr()
        self.createMainScreen()
        self.createTextBox()
        worksheet = self.loadCommunityWorkbook()
        if(self.parseWorkSheet(worksheet)):
            self.displayCommunityMembers(self.text_frame, self.community_members)

    def loadCommunityWorkbook(self):
        FILEOPENOPTIONS = dict(filetypes=[(('Xls','*.xls'),('Xlsx','*.xlsx')),('All files','*.*')])
        community_workbook_path = filedialog.askopenfilename(**FILEOPENOPTIONS)
        if (community_workbook_path != ''):
            wb = load_workbook(filename=community_workbook_path, read_only=True)
            ws = wb.active
            self.loaded = True
        return ws

    def parseWorkSheet(self, ws):
        self.community_members = []
        for row in ws.rows:
            pearson_info = []
            for cell in row[:4]:
                pearson_info.append(cell.value)
            pearson = Pearson(pearson_info)
            self.community_members.append(pearson)
        return True

    def printCommunityMembers(self):
        for community_member in self.community_members:
            community_member.print()

    def displayLoadCommunityFirstWarning(self):
        filewin = tk.Toplevel(self.current_screen)
        label = tk.Label(filewin, text="\nLoad community first")
        button = tk.Button(filewin, text="Ok", command=filewin.destroy)
        filewin.geometry("150x75")
        filewin.title("Warning!")
        label.pack()
        button.pack()

    def displayCommunityMembers(self, textBox, members):
        if self.loaded:
            self.displyMembersInsideTextBox(textBox, members)
        else:
            self.displayLoadCommunityFirstWarning()

    def displyMembersInsideTextBox(self, textBox, members):
        r = 0
        for pearson in members:
            r = r + 1
            c = 0
            for info in pearson.getName():
                c = c + 1
                tk.Label(textBox, text=info, bg = 'white').grid(row=r, column=c, sticky=tk.W, padx=10)
                ttk.Separator(textBox, orient=tk.VERTICAL).grid(column=c+1, row=r, rowspan=len(members))

    def pickNewCircles(self):
        if self.loaded:
            self.circles = [[] for x in range(self.number_of_circles)]
            number_of_members = len(self.community_members)
            picked_indices = []
            while len(picked_indices) < number_of_members:
                for i in range(self.number_of_circles):
                    if len(picked_indices) < number_of_members:
                        index = random.randrange(number_of_members)
                        while (index in picked_indices) and (len(picked_indices) < number_of_members):
                            index = random.randrange(number_of_members)
                        picked_indices.append(index)
                        self.circles[i].append(self.community_members[index])
            self.displayCircles()
        else:
            self.displayLoadCommunityFirstWarning()

    def displayCircles(self):
        self.clearScr()
        self.createMainScreen()
        self.initCircleTextBox()
        for i in range(self.number_of_circles):
            self.displyMembersInsideTextBox(self.circleTextBox[i], self.circles[i])
            circle_label = tk.Label(self.current_screen, text="Circle " + str(i))
            circle_label.pack()
            self.circleTextBox[i].pack()

